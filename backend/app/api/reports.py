from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from typing import Optional, List
from datetime import date, datetime, timedelta
from decimal import Decimal
from fastapi.responses import StreamingResponse


from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User, Case, Aid, AidType, MonthlyAidTransaction, MonthlyAidCycle
from app.schemas.reports import (
    DashboardStatsResponse, AidTypeDistributionResponse,
    MonthlyStatsResponse, TopBeneficiaryResponse,
    CaseStatusResponse, ReportFilters
)

router = APIRouter(tags=["Reports"])

@router.get("/reports/dashboard-stats")
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get main dashboard statistics"""
    
    # Total cases
    total_cases = db.query(Case).filter(Case.is_deleted == False).count()
    active_cases = db.query(Case).filter(Case.is_deleted == False).count()  # All non-deleted are active
    
    # Total aids
    total_aids = db.query(Aid).filter(Aid.is_deleted == False).count()
    total_amount = db.query(func.sum(Aid.amount)).filter(
        Aid.is_deleted == False,
        Aid.amount.isnot(None)
    ).scalar() or 0
    
    # Total beneficiaries (unique cases that received aid)
    total_beneficiaries = db.query(Aid.case_id).distinct().filter(Aid.is_deleted == False).count()
    
    # Total users
    total_users = db.query(User).filter(User.is_deleted == False).count()
    
    # Current month aid
    current_month = date.today().month
    current_year = date.today().year
    monthly_aid_count = db.query(MonthlyAidTransaction).join(
        MonthlyAidCycle
    ).filter(
        MonthlyAidTransaction.is_deleted == False,
        MonthlyAidCycle.year == current_year,
        MonthlyAidCycle.month == current_month
    ).count()
    
    # Pending deliveries for current month
    pending_deliveries = db.query(MonthlyAidTransaction).join(
        MonthlyAidCycle
    ).filter(
        MonthlyAidTransaction.is_deleted == False,
        MonthlyAidTransaction.status == "Pending",
        MonthlyAidCycle.year == current_year,
        MonthlyAidCycle.month == current_month
    ).count()
    
    # Recent activities (last 10)
    recent_activities = []
    
    # Recent cases
    recent_cases = db.query(Case).filter(
        Case.is_deleted == False
    ).order_by(Case.created_at.desc()).limit(5).all()
    
    for case in recent_cases:
        recent_activities.append({
            "action": "create_case",
            "entity": "Case",
            "description": f"تم إضافة حالة جديدة: {case.full_name}",
            "date": case.created_at.isoformat(),
            "user": None
        })
    
    # Recent aids
    recent_aids = db.query(Aid).filter(
        Aid.is_deleted == False
    ).order_by(Aid.created_at.desc()).limit(5).all()
    
    for aid in recent_aids:
        case = db.query(Case).filter(Case.id == aid.case_id).first()
        user = db.query(User).filter(User.id == aid.registered_by_user_id).first()
        recent_activities.append({
            "action": "create_aid",
            "entity": "Aid",
            "description": f"تم تسجيل {aid.aid_type_id} للمستفيد {case.full_name if case else ''}",
            "date": aid.created_at.isoformat(),
            "user": user.full_name if user else None
        })
    
    # Sort by date
    recent_activities.sort(key=lambda x: x["date"], reverse=True)
    recent_activities = recent_activities[:10]
    
    return {
        "total_cases": total_cases,
        "active_cases": active_cases,
        "total_aids": total_aids,
        "total_amount": float(total_amount),
        "total_beneficiaries": total_beneficiaries,
        "total_users": total_users,
        "monthly_aid_cases": monthly_aid_count,
        "pending_deliveries": pending_deliveries,
        "recent_activities": recent_activities
    }

@router.get("/reports/aid-distribution")
def get_aid_distribution(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get aid type distribution for charts"""
    
    query = db.query(
        AidType.name,
        AidType.category,
        func.count(Aid.id).label("count"),
        func.sum(Aid.amount).label("total_amount")
    ).join(Aid, Aid.aid_type_id == AidType.id).filter(
        Aid.is_deleted == False,
        AidType.is_deleted == False
    )
    
    if start_date:
        query = query.filter(Aid.aid_date >= start_date)
    if end_date:
        query = query.filter(Aid.aid_date <= end_date)
    
    results = query.group_by(AidType.id).all()
    
    return [
        {
            "name": r.name,
            "category": r.category,
            "count": r.count,
            "total_amount": float(r.total_amount) if r.total_amount else 0
        }
        for r in results
    ]

@router.get("/reports/monthly-stats")
def get_monthly_stats(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get monthly statistics for charts"""
    
    if not year:
        year = date.today().year
    
    monthly_data = []
    
    for month in range(1, 13):
        # Aid count and amount for this month
        aid_stats = db.query(
            func.count(Aid.id).label("count"),
            func.sum(Aid.amount).label("amount")
        ).filter(
            Aid.is_deleted == False,
            extract('year', Aid.aid_date) == year,
            extract('month', Aid.aid_date) == month
        ).first()
        
        # Monthly aid cases for this month
        monthly_aid = db.query(MonthlyAidTransaction).join(
            MonthlyAidCycle
        ).filter(
            MonthlyAidTransaction.is_deleted == False,
            MonthlyAidCycle.year == year,
            MonthlyAidCycle.month == month
        ).count()
        
        monthly_data.append({
            "month": month,
            "aid_count": aid_stats.count or 0,
            "aid_amount": float(aid_stats.amount) if aid_stats.amount else 0,
            "monthly_aid_cases": monthly_aid
        })
    
    return monthly_data

@router.get("/reports/top-beneficiaries")
def get_top_beneficiaries(
    limit: int = Query(10, ge=1, le=50),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get top beneficiaries by aid amount and count"""
    
    query = db.query(
        Case.id,
        Case.full_name,
        Case.file_number,
        func.count(Aid.id).label("aid_count"),
        func.sum(Aid.amount).label("total_amount")
    ).join(Aid, Aid.case_id == Case.id).filter(
        Aid.is_deleted == False,
        Case.is_deleted == False
    )
    
    if start_date:
        query = query.filter(Aid.aid_date >= start_date)
    if end_date:
        query = query.filter(Aid.aid_date <= end_date)
    
    results = query.group_by(Case.id).order_by(
        func.sum(Aid.amount).desc()
    ).limit(limit).all()
    
    return [
        {
            "id": r.id,
            "full_name": r.full_name,
            "file_number": r.file_number,
            "aid_count": r.aid_count,
            "total_amount": float(r.total_amount) if r.total_amount else 0
        }
        for r in results
    ]

@router.get("/reports/case-status")
def get_case_status_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get case status statistics"""
    
    # Active cases (not deleted)
    active_cases = db.query(Case).filter(Case.is_deleted == False).count()
    
    # Cases with monthly aid
    monthly_aid_cases = db.query(Case).filter(
        Case.is_monthly_aid == True,
        Case.is_deleted == False
    ).count()
    
    # Cases that received aid
    cases_with_aid = db.query(Aid.case_id).distinct().filter(
        Aid.is_deleted == False
    ).count()
    
    # Cases by city
    city_stats = db.query(
        Case.city,
        func.count(Case.id).label("count")
    ).filter(Case.is_deleted == False).group_by(Case.city).all()
    
    return {
        "active_cases": active_cases,
        "monthly_aid_cases": monthly_aid_cases,
        "cases_with_aid": cases_with_aid,
        "cases_by_city": [{"city": r.city, "count": r.count} for r in city_stats]
    }

@router.get("/reports/export-excel")
def export_reports_to_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    report_type: str = "comprehensive",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export reports to Excel"""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from urllib.parse import quote
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # ============ Sheet 1: Dashboard Statistics ============
    if report_type in ["comprehensive"]:
        ws1 = wb.create_sheet("الإحصائيات العامة")
        
        # Title
        ws1.merge_cells('A1:D1')
        ws1['A1'] = 'الإحصائيات العامة'
        ws1['A1'].font = Font(bold=True, size=16)
        ws1['A1'].alignment = Alignment(horizontal="center")
        
        # Get stats
        stats = get_dashboard_stats(db, current_user)
        
        # Add data
        data_rows = [
            ('إجمالي الحالات', stats.get('total_cases', 0)),
            ('الحالات النشطة', stats.get('active_cases', 0)),
            ('إجمالي المساعدات', stats.get('total_aids', 0)),
            ('إجمالي المبلغ (ج.م)', stats.get('total_amount', 0)),
            ('إجمالي المستفيدين', stats.get('total_beneficiaries', 0)),
            ('إجمالي المستخدمين', stats.get('total_users', 0)),
            ('حالات المساعدات الشهرية', stats.get('monthly_aid_cases', 0)),
            ('تسليمات معلقة', stats.get('pending_deliveries', 0)),
        ]
        
        for idx, (label, value) in enumerate(data_rows, start=3):
            ws1.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws1.cell(row=idx, column=2, value=value)
        
        # Adjust column widths
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
    
    # ============ Sheet 2: Aid Distribution ============
    if report_type in ["comprehensive", "aids"]:
        ws2 = wb.create_sheet("توزيع المساعدات")
        
        # Title
        ws2.merge_cells('A1:D1')
        ws2['A1'] = 'توزيع المساعدات حسب النوع'
        ws2['A1'].font = Font(bold=True, size=16)
        ws2['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['نوع المساعدة', 'الفئة', 'العدد', 'إجمالي المبلغ (ج.م)']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Get distribution data
        distribution = get_aid_distribution(start_date, end_date, db, current_user)
        
        for idx, item in enumerate(distribution, start=4):
            ws2.cell(row=idx, column=1, value=item.get('name', ''))
            ws2.cell(row=idx, column=2, value=item.get('category', ''))
            ws2.cell(row=idx, column=3, value=item.get('count', 0))
            ws2.cell(row=idx, column=4, value=float(item.get('total_amount', 0)))
        
        # Adjust column widths
        for col in range(1, 5):
            ws2.column_dimensions[chr(64 + col)].width = 20
    
    # ============ Sheet 3: Monthly Statistics ============
    if report_type in ["comprehensive"]:
        ws3 = wb.create_sheet("الإحصائيات الشهرية")
        
        # Title
        ws3.merge_cells('A1:D1')
        ws3['A1'] = 'الإحصائيات الشهرية'
        ws3['A1'].font = Font(bold=True, size=16)
        ws3['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['الشهر', 'عدد المساعدات', 'المبلغ (ج.م)', 'حالات المساعدات الشهرية']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Get monthly data
        monthly_stats = get_monthly_stats(None, db, current_user)
        month_names = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
        
        for idx, item in enumerate(monthly_stats, start=4):
            ws3.cell(row=idx, column=1, value=month_names[item.get('month', 1) - 1])
            ws3.cell(row=idx, column=2, value=item.get('aid_count', 0))
            ws3.cell(row=idx, column=3, value=float(item.get('aid_amount', 0)))
            ws3.cell(row=idx, column=4, value=item.get('monthly_aid_cases', 0))
        
        # Adjust column widths
        for col in range(1, 5):
            ws3.column_dimensions[chr(64 + col)].width = 20
    
    # ============ Sheet 4: Top Beneficiaries ============
    if report_type in ["comprehensive", "beneficiaries"]:
        ws4 = wb.create_sheet("أكثر المستفيدين")
        
        # Title
        ws4.merge_cells('A1:E1')
        ws4['A1'] = 'أكثر المستفيدين'
        ws4['A1'].font = Font(bold=True, size=16)
        ws4['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['#', 'رقم الملف', 'اسم المستفيد', 'عدد المساعدات', 'إجمالي المبلغ (ج.م)']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Get beneficiaries data
        beneficiaries = get_top_beneficiaries(50, start_date, end_date, db, current_user)
        
        for idx, item in enumerate(beneficiaries, start=4):
            ws4.cell(row=idx, column=1, value=idx - 3)
            ws4.cell(row=idx, column=2, value=item.get('file_number', ''))
            ws4.cell(row=idx, column=3, value=item.get('full_name', ''))
            ws4.cell(row=idx, column=4, value=item.get('aid_count', 0))
            ws4.cell(row=idx, column=5, value=float(item.get('total_amount', 0)))
        
        # Adjust column widths
        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 25
        ws4.column_dimensions['D'].width = 20
        ws4.column_dimensions['E'].width = 20
    
    # ============ Sheet 5: Case Status ============
    if report_type in ["comprehensive"]:
        ws5 = wb.create_sheet("حالة الحالات")
        
        # Title
        ws5.merge_cells('A1:B1')
        ws5['A1'] = 'حالة الحالات'
        ws5['A1'].font = Font(bold=True, size=16)
        ws5['A1'].alignment = Alignment(horizontal="center")
        
        # Get case status
        case_status = get_case_status_stats(db, current_user)
        
        # Add data
        status_rows = [
            ('الحالات النشطة', case_status.get('active_cases', 0)),
            ('حالات المساعدات الشهرية', case_status.get('monthly_aid_cases', 0)),
            ('حالات استلمت مساعدات', case_status.get('cases_with_aid', 0)),
        ]
        
        for idx, (label, value) in enumerate(status_rows, start=3):
            ws5.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws5.cell(row=idx, column=2, value=value)
        
        # Cases by city
        ws5.cell(row=6, column=1, value="توزيع الحالات حسب المدينة").font = Font(bold=True, size=12)
        
        cities = case_status.get('cases_by_city', [])
        for idx, city_data in enumerate(cities, start=8):
            ws5.cell(row=idx, column=1, value=city_data.get('city', ''))
            ws5.cell(row=idx, column=2, value=city_data.get('count', 0))
        
        # Adjust column widths
        ws5.column_dimensions['A'].width = 25
        ws5.column_dimensions['B'].width = 20
    
    # ============ Sheet 6: Recent Activities ============
    if report_type in ["comprehensive"]:
        ws6 = wb.create_sheet("آخر الأنشطة")
        
        # Title
        ws6.merge_cells('A1:C1')
        ws6['A1'] = 'آخر الأنشطة'
        ws6['A1'].font = Font(bold=True, size=16)
        ws6['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['التاريخ', 'النشاط', 'بواسطة']
        for col, header in enumerate(headers, 1):
            cell = ws6.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Get stats which includes recent activities
        stats = get_dashboard_stats(db, current_user)
        activities = stats.get('recent_activities', [])
        
        for idx, activity in enumerate(activities, start=4):
            ws6.cell(row=idx, column=1, value=activity.get('date', ''))
            ws6.cell(row=idx, column=2, value=activity.get('description', ''))
            ws6.cell(row=idx, column=3, value=activity.get('user', 'نظام'))
        
        # Adjust column widths
        ws6.column_dimensions['A'].width = 20
        ws6.column_dimensions['B'].width = 40
        ws6.column_dimensions['C'].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"report_{report_type}_{date.today()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )   
    
    
    
    
    
    