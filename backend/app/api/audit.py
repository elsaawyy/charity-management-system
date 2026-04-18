from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import Optional, List
from datetime import date, datetime

from app.core.database import get_db
from app.core.security import get_current_user, require_admin
from app.models.user import AuditLog, User
from app.schemas.audit import AuditLogResponse, AuditLogListResponse

router = APIRouter(tags=["Audit"])

@router.get("/audit-logs", response_model=AuditLogListResponse)
def list_audit_logs(
    page: int = Query(1, ge=1),
    size: int = Query(10, ge=1, le=100),
    user_id: Optional[int] = None,
    action_type: Optional[str] = None,
    entity_name: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    query = db.query(AuditLog)
    
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    
    if action_type:
        query = query.filter(AuditLog.action_type == action_type)
    
    if entity_name:
        query = query.filter(AuditLog.entity_name == entity_name)
    
    if start_date:
        query = query.filter(func.date(AuditLog.action_date) >= start_date)
    
    if end_date:
        query = query.filter(func.date(AuditLog.action_date) <= end_date)
    
    if search:
        query = query.filter(
            or_(
                AuditLog.description.ilike(f"%{search}%"),
                AuditLog.entity_name.ilike(f"%{search}%")
            )
        )
    
    total = query.count()
    items = query.order_by(AuditLog.action_date.desc()).offset((page - 1) * size).limit(size).all()
    
    # Enrich with user details
    result_items = []
    for log in items:
        user = db.query(User).filter(User.id == log.user_id).first() if log.user_id else None
        result_items.append({
            "id": log.id,
            "user_id": log.user_id,
            "username": user.username if user else None,
            "user_full_name": user.full_name if user else None,
            "action_type": log.action_type,
            "entity_name": log.entity_name,
            "entity_id": log.entity_id,
            "old_values": log.old_values,
            "new_values": log.new_values,
            "description": log.description,
            "ip_address": log.ip_address,
            "action_date": log.action_date
        })
    
    return {
        "items": result_items,
        "total": total,
        "page": page,
        "size": size,
        "pages": (total + size - 1) // size
    }

@router.get("/audit-logs/actions")
def get_action_types(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    actions = db.query(AuditLog.action_type).distinct().all()
    return [a[0] for a in actions]

@router.get("/audit-logs/export")
def export_audit_logs(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from urllib.parse import quote
    from fastapi.responses import StreamingResponse
    
    query = db.query(AuditLog).order_by(AuditLog.action_date.desc())
    
    if start_date:
        query = query.filter(func.date(AuditLog.action_date) >= start_date)
    if end_date:
        query = query.filter(func.date(AuditLog.action_date) <= end_date)
    
    logs = query.limit(10000).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    
    # Headers
    headers = ['التاريخ', 'المستخدم', 'نوع الإجراء', 'الكيان', 'معرف الكيان', 'الوصف', 'عنوان IP']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for idx, log in enumerate(logs, start=2):
        user = db.query(User).filter(User.id == log.user_id).first() if log.user_id else None
        ws.cell(row=idx, column=1, value=log.action_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=idx, column=2, value=user.full_name if user else 'System')
        ws.cell(row=idx, column=3, value=log.action_type)
        ws.cell(row=idx, column=4, value=log.entity_name)
        ws.cell(row=idx, column=5, value=log.entity_id)
        ws.cell(row=idx, column=6, value=log.description)
        ws.cell(row=idx, column=7, value=log.ip_address)
    
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('audit_logs.xlsx')}"}
    )